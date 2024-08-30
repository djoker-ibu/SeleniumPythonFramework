import openpyxl


class HomePageData:
    """
    test_HomePae_data = [{"firstname": "Ibu", "email": "djoker@mail.com", "gender": "Male"},
                         {"firstname": "Bae", "email": "djoking@mail.com", "gender": "Female"}]
    """

    @staticmethod  # To call this method with creating an object from class we declare staticmethod
    # Also self is not declared cause it is declared as staticmethod
    def getTestData(test_case_name):
        dict = {}
        book = openpyxl.load_workbook(
            "F:\PythonProjects\SeleniumPythonFramework\TestData\HomePageData.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(1, sheet.max_column + 1):
                    # dict["firstname"] = "Ibu"
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        # print(dict)
        return [dict]
